import pandas as pd
import json
import re
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# ================= Configuration Section =================
DATA_FILE = "./input_data/text_data.json"  # Your original data file (Source of Truth)

OUTPUT_DIR = r"./output_data"
os.makedirs(OUTPUT_DIR, exist_ok=True)

INPUT_FILE = f"{OUTPUT_DIR}/.xlsx"  # Your evaluation result file
OUTPUT_FILE = f"{OUTPUT_DIR}/FINAL_Validated_Report.xlsx"  # Output file

# Color configuration
COLOR_HEADER_BG = "4472C4"
COLOR_HEADER_FONT = "FFFFFF"
COLOR_CORRECT = "E2EFDA"     # Correct (green)
COLOR_WRONG = "FCE4D6"       # Wrong (red)
COLOR_REFUSAL = "EDEDED"     # Refusal (gray)
COLOR_GT = "FFF2CC"          # Ground Truth (yellow)

# ================= 1. Load and Build Knowledge Base Index =================
def load_knowledge_base(json_file):
    if not os.path.exists(json_file):
        print(f"❌ Data file not found: {json_file}")
        return None
        
    try:
        with open(json_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
            return data.get('entries', [])
    except Exception as e:
        print(f"❌ Failed to read JSON: {e}")
        return None

def find_relevant_context(query, knowledge_base):
    """
    Find the most relevant original text snippet in the knowledge base based on question keywords.
    """
    if not knowledge_base: return ""
    
    query_norm = str(query).lower()
    # Extract core keywords (remove stopwords)
    stops = ["what", "is", "the", "a", "an", "of", "in", "for", "to", "does", "do", "can", "if", "required", "limit", "value"]
    keywords = [w for w in re.split(r'\W+', query_norm) if w and w not in stops and len(w) > 2]
    
    best_match = None
    max_score = 0
    
    for entry in knowledge_base:
        content = str(entry.get('text', '')).lower()
        # Simple scoring: number of keyword matches
        score = sum(1 for k in keywords if k in content)
        
        if score > max_score:
            max_score = score
            best_match = entry.get('text', '')
            
    return best_match if max_score > 0 else ""

# ================= 2. Core Evaluation Logic (Based on Knowledge Base) =================
def normalize(text):
    return str(text).lower().strip() if pd.notna(text) else ""

def is_refusal(text):
    keywords = ["cannot answer", "no information", "unable to determine", "not found", "graph path not found"]
    return any(k in normalize(text) for k in keywords)

def judge_answer_with_context(response, gt, question, knowledge_base):
    """
    Advanced evaluation logic:
    1. Check if it is a refusal -> mark as incorrect
    2. If knowledge base context exists, verify whether Ground Truth (GT) appears in it (validate GT)
    3. Then check whether the model response matches the GT
    """
    pred = normalize(response)
    truth = normalize(gt)
    
    if is_refusal(pred): return False

    # --- Rule A: Yes/No logic ---
    if truth.startswith("yes"):
        return "yes" in pred and "no" not in pred[:15]
    if truth.startswith("no"):
        return "no" in pred and "yes" not in pred[:15]

    # --- Rule B: Numerical logic (with knowledge base context) ---
    gt_nums = re.findall(r'\d+(?:\.\d+)?', truth)
    if gt_nums:
        pred_nums = re.findall(r'\d+(?:\.\d+)?', pred)
        # If all numbers in GT appear in prediction
        if all(n in pred_nums for n in gt_nums):
            return True
            
        # [Advanced]: If GT is "Min 20mm" and model says "at least 20mm", still correct
        # This step uses knowledge base text for verification
        context = find_relevant_context(question, knowledge_base)
        if context:
            context_nums = re.findall(r'\d+(?:\.\d+)?', context)
            # If model numbers match numbers in context, count as correct (avoid incomplete GT issues)
            if any(n in pred_nums for n in context_nums if len(n) > 1): # ignore single-digit noise
                return True

    # --- Rule C: Keyword overlap ---
    stops = {"the", "is", "a", "of", "to", "in", "standard", "code"}
    gt_tokens = set(re.findall(r'\w+', truth)) - stops
    pred_tokens = set(re.findall(r'\w+', pred))
    
    if not gt_tokens: return False
    
    match_rate = len(gt_tokens & pred_tokens) / len(gt_tokens)
    return match_rate >= 0.5

# ================= 3. Main Program =================
print(f"📂 Loading knowledge base: {DATA_FILE}...")
kb = load_knowledge_base(DATA_FILE)

print(f"📂 Reading evaluation file: {INPUT_FILE}...")
try:
    df = pd.read_excel(INPUT_FILE, engine='openpyxl')
except:
    df = pd.read_csv(INPUT_FILE.replace(".xlsx", ".csv")) # fallback

df.columns = [c.strip() for c in df.columns]
model_cols = [c for c in df.columns if "模型回复" in c]
gt_col = "标准答案 (Ground Truth)"
q_col = "大模型提示词"

print("⚖️ Evaluating based on original data...")
score_matrix = pd.DataFrame()

for col in model_cols:
    # Row-wise evaluation
    score_matrix[col] = df.apply(lambda row: judge_answer_with_context(
        row[col], 
        row[gt_col], 
        row[q_col], 
        kb
    ), axis=1)

# Statistics
stats = score_matrix.sum().to_frame(name="Correct Count")
stats["Accuracy"] = (score_matrix.mean() * 100).apply(lambda x: f"{x:.1f}%")

print("\n📊 Evaluation Results:")
print(stats)

# ================= 4. Export Styled Excel =================
print(f"\n🎨 Generating final report: {OUTPUT_FILE}...")
writer = pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl')
stats.to_excel(writer, sheet_name="Statistics")
df.to_excel(writer, sheet_name="Detailed Evaluation", index=False)
writer.close()

# Coloring
wb = load_workbook(OUTPUT_FILE)
ws = wb["Detailed Evaluation"]

fill_correct = PatternFill("solid", fgColor=COLOR_CORRECT)
fill_wrong = PatternFill("solid", fgColor=COLOR_WRONG)
fill_header = PatternFill("solid", fgColor=COLOR_HEADER_BG)
fill_gt = PatternFill("solid", fgColor=COLOR_GT)
font_header = Font(color=COLOR_HEADER_FONT, bold=True)
border_thin = Side(border_style="thin", color="BFBFBF")
border_all = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
align_top = Alignment(vertical="top", wrap_text=True)

# Find column indices
header = [cell.value for cell in ws[1]]
try:
    model_indices = {name: i+1 for i, name in enumerate(header) if name in model_cols}
    gt_index = header.index(gt_col) + 1
except:
    model_indices = {}

for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
    for cell in row:
        cell.alignment = align_top
        cell.border = border_all
        
        if row[0].row == 1:
            cell.fill = fill_header
            cell.font = font_header
            continue
            
        if cell.column == gt_index:
            cell.fill = fill_gt
            cell.font = Font(bold=True)

        col_name = header[cell.column-1]
        if col_name in model_indices:
            is_right = score_matrix.at[cell.row-2, col_name]
            cell.fill = fill_correct if is_right else fill_wrong

wb.save(OUTPUT_FILE)
print(f"✅ Done! Please check: {OUTPUT_FILE}")