import os
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ================= Configuration Section =================

# Output file save path (recommended to save outside the folder or use a different name)
OUTPUT_DIR = r"./output_data"
os.makedirs(OUTPUT_DIR, exist_ok=True)

# Input file
INPUT_PATH = f"{OUTPUT_DIR}/Graph_RAG_Correct_Only.xlsx"
OUTPUT_PATH = f"{OUTPUT_DIR}/Benchmark_Final_Analysis.xlsx"

# --- Style definitions ---
# Color configuration
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
GREEN_FONT = Font(color="006100")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
RED_FONT = Font(color="9C0006")

# Header style
HEADER_FILL = PatternFill(start_color="E9E9E9", end_color="E9E9E9", fill_type="solid")
HEADER_FONT = Font(bold=True, name='Microsoft YaHei', size=11)
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Border style
THIN_BORDER = Border(left=Side(style='thin', color='B2B2B2'),
                     right=Side(style='thin', color='B2B2B2'),
                     top=Side(style='thin', color='B2B2B2'),
                     bottom=Side(style='thin', color='B2B2B2'))

# Column width limits (characters)
MAX_WIDTH_TEXT = 60    # Question/Answer columns max width
MAX_WIDTH_RES = 15     # Result columns max width
MAX_WIDTH_ID = 8       # ID column max width

def beautify_and_analyze():
    print(f"🚀 Starting file processing: {INPUT_PATH}")
    
    # 1. Read data using Pandas for statistics
    try:
        df = pd.read_excel(INPUT_PATH)
    except FileNotFoundError:
        print(f"❌ File not found, please check path: {INPUT_PATH}")
        return
    except Exception as e:
        print(f"❌ Error reading file: {e}")
        return

    # Identify all Res columns
    res_columns = [col for col in df.columns if "Res" in col]
    print(f"📊 Detected {len(res_columns)} model result columns: {res_columns}")

    # Compute statistics
    stats = {}
    total_rows = len(df)
    for col in res_columns:
        # Count entries containing "Correct"
        correct_num = df[col].astype(str).str.contains("Correct", na=False).sum()
        ratio = (correct_num / total_rows * 100) if total_rows > 0 else 0
        stats[col] = {"Correct": correct_num, "Ratio": ratio}

    # 2. Use OpenPyXL for formatting
    print("🎨 Applying conditional formatting and layout...")
    wb = openpyxl.load_workbook(INPUT_PATH)
    ws = wb.active
    max_row = ws.max_row
    max_col = ws.max_column

    # Get header mapping
    headers = [cell.value for cell in ws[1]]
    res_col_indices = [i + 1 for i, h in enumerate(headers) if "Res" in str(h)]
    
    # --- Iterate through all cells ---
    for r_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=max_row, max_col=max_col), start=2):
        for c_idx, cell in enumerate(row, start=1):
            cell.border = THIN_BORDER
            
            # Alignment
            if c_idx == 1 or (headers[c_idx-1] and "ID" in str(headers[c_idx-1])): 
                cell.alignment = Alignment(horizontal='center', vertical='top')
            else: 
                cell.alignment = Alignment(vertical='top', wrap_text=True)

            # Conditional coloring
            if c_idx in res_col_indices and cell.value:
                val = str(cell.value)
                if "Correct" in val:
                    cell.fill = GREEN_FILL
                    cell.font = GREEN_FONT
                elif "Incorrect" in val:
                    cell.fill = RED_FILL
                    cell.font = RED_FONT

    # --- Header styling ---
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    # --- Smart column width adjustment ---
    print("📏 Adjusting column widths...")
    for i, col_cells in enumerate(ws.columns, start=1):
        col_letter = get_column_letter(i)
        header_val = str(col_cells[0].value) if col_cells[0].value else ""

        # Set target width
        if "ID" in header_val: target = MAX_WIDTH_ID
        elif "Res" in header_val or "Truth" in header_val: target = MAX_WIDTH_RES
        else: target = MAX_WIDTH_TEXT

        # Compute content length
        max_len = 0
        for cell in col_cells[:50]: # sample first 50 rows
            if cell.value:
                try:
                    # Fix: add errors='ignore' to prevent encoding issues
                    curr_len = len(str(cell.value).encode('gbk', errors='ignore')) * 0.9
                except:
                    # fallback estimation
                    curr_len = len(str(cell.value)) * 1.0
                max_len = max(max_len, curr_len)
        
        final_width = min(max_len + 2, target)
        final_width = max(final_width, 10) # minimum width
        ws.column_dimensions[col_letter].width = final_width

    # --- Add statistics summary table ---
    print("📈 Writing statistics summary...")
    start_row = max_row + 3
    
    # Title
    ws.cell(row=start_row, column=2, value="📊 Model Performance Summary").font = Font(bold=True, size=12)
    
    # Summary headers
    headers_stats = ["Model Name", "Correct Count", "Accuracy (%)", "Total Questions"]
    for i, h in enumerate(headers_stats):
        c = ws.cell(row=start_row + 1, column=2 + i, value=h)
        c.fill = HEADER_FILL
        c.font = Font(bold=True)
        c.border = THIN_BORDER
        c.alignment = Alignment(horizontal='center')

    # Summary data rows
    for i, (model_name, stat) in enumerate(stats.items()):
        r = start_row + 2 + i
        
        # Model name
        c1 = ws.cell(row=r, column=2, value=model_name.replace("_Res", ""))
        c1.border = THIN_BORDER
        c1.alignment = Alignment(horizontal='left')
        
        # Correct count
        c2 = ws.cell(row=r, column=3, value=stat['Correct'])
        c2.border = THIN_BORDER
        c2.alignment = Alignment(horizontal='center')
        
        # Accuracy
        c3 = ws.cell(row=r, column=4, value=f"{stat['Ratio']:.2f}%")
        c3.border = THIN_BORDER
        c3.alignment = Alignment(horizontal='center')
        if stat['Ratio'] >= 99.9: 
            c3.font = GREEN_FONT
            c3.fill = GREEN_FILL
        
        # Total
        c4 = ws.cell(row=r, column=5, value=total_rows)
        c4.border = THIN_BORDER
        c4.alignment = Alignment(horizontal='center')

    # 3. Save file
    try:
        wb.save(OUTPUT_PATH)
        print(f"\n✅ Completed! File saved to:\n{OUTPUT_PATH}")
    except PermissionError:
        print(f"\n❌ Save failed! Please close the Excel file: {OUTPUT_PATH}")

if __name__ == "__main__":
    beautify_and_analyze()