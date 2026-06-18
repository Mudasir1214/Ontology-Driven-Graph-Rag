from dotenv import load_dotenv
import pandas as pd
import time
import os
import json
import re
from dotenv import load_dotenv
from tqdm import tqdm
from openai import OpenAI
from neo4j import GraphDatabase
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Font

# Load environment variables from .env file
load_dotenv()  

# ================= 1. Core Configuration =================

API_KEY = os.getenv("DEEPSEEK_API_KEY")
BASE_URL = os.getenv("DEEPSEEK_BASE_URL")
NEO4J_URI = os.getenv("NEO4J_URI")
NEO4J_AUTH = (os.getenv("NEO4J_USER_NAME"), os.getenv("NEO4J_PASSWORD"))

# Output file save path (recommended to save outside the folder or use a different name)
OUTPUT_DIR = r"./output_data"
os.makedirs(OUTPUT_DIR, exist_ok=True)

OUTPUT_FILE = f"{OUTPUT_DIR}/Benchmark_Final_600.xlsx"
TEXT_DATA_FILE = f"./input_data/text_data.json" 

# ================= 2. Question Merge Section =================

QUESTIONS_DATA = [
    # --- PART 1: Existing Success Cases (51题) ---
    {"id": 1, "q": "According to Fire Code 2022, which standard applies to 'Automatic actuating device' cables if the diameter is 25mm?", "gt": "BS 7846"},
    {"id": 2, "q": "According to Fire Code 2022, which standard applies to 'Automatic actuating device' cables if the diameter is 15mm?", "gt": "BS 7629-1"},
    {"id": 3, "q": "Does the 2022 Fire Code allow BS 6207 for 'Emergency generator' power supply?", "gt": "No (Removed in 2022)"},
    {"id": 8, "q": "Is 'Ductile iron pipe' mandatory for a 90mm underground fire service pipe?", "gt": "Yes (>=80mm rule applies)"},
    {"id": 11, "q": "Is a 40mm high 'HV warning notice' compliant with the 2020 Regulations?", "gt": "No (Minimum 50mm required)"},
    {"id": 12, "q": "Is a 40mm high 'HV warning notice' compliant with the 2015 Regulations?", "gt": "Yes (Minimum 30mm required)"},
    {"id": 15, "q": "Can I install a 'Ceiling fan' at a height of 2.3 meters in 2022?", "gt": "No (Minimum 2400mm required)"},
    {"id": 16, "q": "Can I install a 'Ceiling fan' at a height of 2.7 meters in 2022?", "gt": "No (Maximum 2600mm required)"},
    {"id": 19, "q": "Can the 'CCMS' off time be set to 97%?", "gt": "No (Maximum 95% allowed)"},
    {"id": 20, "q": "Can the 'CCMS' off time be set to 3%?", "gt": "No (Minimum 5% allowed)"},
    {"id": 21, "q": "Is a 5mm clearance sufficient for 'Restrained spring mount' (Type B)?", "gt": "No (Minimum 10mm required)"},
    {"id": 26, "q": "Can a screw shank protrude by 5 threads on a 'Saddle'?", "gt": "No (Maximum 3 threads allowed)"},
    {"id": 27, "q": "Is intermediate fixing required for a 100mm 'Saddle'?", "gt": "No (Only required for >150mm)"},
    {"id": 29, "q": "Is a starting current of 3 times full load allowed for 'Auto-transformer'?", "gt": "No (Maximum 2.5 times allowed)"},
    {"id": 30, "q": "Is a 20 degree swing capability sufficient for 'Spring hanger' (Type E)?", "gt": "No (30-35 degrees required)"},
    {"id": 31, "q": "Does the 2012 Fire Code mention 'BS 8491' for emergency generators?", "gt": "No (Only introduced in later/other versions)"},
    {"id": 32, "q": "Does the 2022 Fire Code mention 'BS 8491' for automatic actuating devices?", "gt": "Yes (Listed in Item 2)"},
    {"id": 36, "q": "Is 'Polyethylene' a permitted material for 'Acoustic duct lag'?", "gt": "No specific mention / No"},
    {"id": 39, "q": "Does 'Table 4.1' in the 2020 code apply to copper conductors?", "gt": "Yes (or implied context)"},
    {"id": 42, "q": "Is a 'Pump foundation' mass of 1.5 times pump mass acceptable?", "gt": "No (Minimum 2.5 times required)"},
    {"id": 43, "q": "Is a 'Pump foundation' mass of 3.0 times pump mass acceptable?", "gt": "Yes (>= 2.5 times pump mass)"},
    {"id": 44, "q": "Is 'Sand cement screed' acceptable for pump foundation finishing?", "gt": "No (HAC mortar required implied)"},
    {"id": 45, "q": "Is 'HAC mortar' acceptable for pump foundation finishing?", "gt": "Yes"},
    {"id": 46, "q": "Which code governs the 'Testing of fire service installation'?", "gt": "Fire Service"},
    {"id": 47, "q": "Does the 'Plumbing Specification' cover 'Fire Service' pipes?", "gt": "No (Separate documents)"},
    {"id": 51, "q": "List all acceptable standards for 'Automatic actuating device' cables in 2022.", "gt": "BS 6387 CWZ, BS EN 60702, BS 8491, BS 7629-1, BS 7846"},
    {"id": 52, "q": "Can I use a 20W bulb to calculate 'Lighting load' in 2020?", "gt": "No (Minimum 60W applies)"},
    {"id": 56, "q": "Is a 'Minimum bending radius' of 6D acceptable for 'Armoured cable'?", "gt": "No (Usually 8D or specific value)"},
    {"id": 59, "q": "Is 'Schedule 40' steel pipe permitted for fire sprinklers?", "gt": "Yes (Standard spec)"},
    {"id": 60, "q": "Is 'Schedule 10' steel pipe permitted for fire sprinklers?", "gt": "Yes (Often allowed for larger sizes)"},
    {"id": 63, "q": "Is a 'Ductile iron' pipe diameter of 20mm common for underground fire service?", "gt": "No (Typically larger sizes)"},
    {"id": 64, "q": "Is a 'Ductile iron' pipe diameter of 100mm permitted for underground fire service?", "gt": "Yes (Min 25mm)"},
    {"id": 67, "q": "Can 'PVC conduit' be used for 'Emergency lighting' wiring?", "gt": "No (Fire resistant req usually)"},
    {"id": 68, "q": "Is 'Steel conduit' mandatory for 'Emergency lighting' wiring?", "gt": "Yes (Typically for mechanical protection/fire)"},
    {"id": 71, "q": "What is the minimum depth for an 'Underground cable' trench?", "gt": "450mm / 750mm (Context dependent)"},
    {"id": 72, "q": "Is a trench depth of 300mm acceptable for HV cables?", "gt": "No"},
    {"id": 75, "q": "Does 'Appendix A' of the Electrical Code list 'Cable capacities'?", "gt": "Yes/No (Context dependent)"},
    {"id": 76, "q": "Is 'BS 7671' the primary standard for Hong Kong wiring?", "gt": "No (BS 7671 is UK; HK has own CoP based on it)"},
    {"id": 77, "q": "What material is required for 'Pump foundation' finishing?", "gt": "HAC mortar"},
    {"id": 78, "q": "Is 'Grade 304' stainless steel acceptable for 'Seawater' applications?", "gt": "No (Grade 316 usually required)"},
    {"id": 79, "q": "Is 'Grade 316' stainless steel acceptable for 'Seawater' applications?", "gt": "Yes"},
    {"id": 82, "q": "Is 'Class O' insulation required for 'Chilled water pipes'?", "gt": "Yes"},
    {"id": 83, "q": "Is 'Polystyrene' insulation permitted for 'Chilled water pipes'?", "gt": "No (Fire risk)"},
    {"id": 86, "q": "What is the maximum support spacing for a 100mm 'Cable tray'?", "gt": "1.2m / 1.5m (Context dependent)"},
    {"id": 87, "q": "Is a support spacing of 2 meters acceptable for a 100mm 'Cable tray'?", "gt": "No"},
    {"id": 90, "q": "What is the minimum height for a 'Ceiling fan'?", "gt": "2.4m"},
    {"id": 91, "q": "What is the maximum height for a 'Ceiling fan'?", "gt": "2.6m"},
    {"id": 94, "q": "Is 'BS 546' the standard for 'Socket outlets'?", "gt": "Yes (Old style) / No (BS 1363 common)"},
    {"id": 95, "q": "Is 'BS 1363' the standard for '13A Socket outlets'?", "gt": "Yes"},
    {"id": 98, "q": "Does the 'General Specification' include 'Lift' installation?", "gt": "Yes/No (Usually separate)"},
    {"id": 99, "q": "Is 'Escalator' installation covered in the 'Electrical Code'?", "gt": "No (Separate Lift/Escalator code)"},

    # --- PART 2: New Structural/Meta Questions (100题) ---
    {"id": 101, "q": "Does the 'Mechanical Installation' chapter include a section on 'Vibration Isolation'?", "gt": "Yes"},
    {"id": 102, "q": "List all section titles defined under 'Chapter 6: Air Conditioning'.", "gt": "List"},
    {"id": 103, "q": "Which chapter governs the specifications for 'Pipework' in the 2017 Plumbing Code?", "gt": "Chapter Name"},
    {"id": 104, "q": "Does Section 5.1 of the Fire Code contain any sub-clauses?", "gt": "Yes"},
    {"id": 105, "q": "What is the parent chapter for the rule defining 'CCMS' off-time?", "gt": "Chapter 6/CCMS"},
    {"id": 106, "q": "Are 'Electrical Safety' rules defined in the same document as 'Fire Service' rules?", "gt": "No"},
    {"id": 107, "q": "Does the 2022 General Specification have a dedicated Appendices section?", "gt": "Yes"},
    {"id": 108, "q": "How many subsections are there in Chapter 7 of the Electrical Code?", "gt": "Count"},
    {"id": 109, "q": "Is 'Ventilation' covered under Chapter 5 or Chapter 6?", "gt": "Chapter 6"},
    {"id": 110, "q": "Does the document structure distinguish between 'HVAC' and 'Plumbing' chapters explicitly?", "gt": "Yes"},
    {"id": 111, "q": "What is the title of Chapter 8 in the Fire Code?", "gt": "Title"},
    {"id": 112, "q": "Does 'Section 3.2' belong to the 'General Specification' or 'Particular Specification'?", "gt": "General"},
    {"id": 113, "q": "List the hierarchy of sections for 'Cable Installation'.", "gt": "Path"},
    {"id": 114, "q": "Is 'Testing and Commissioning' a top-level chapter?", "gt": "Yes"},
    {"id": 115, "q": "Does the graph show a 'part-of' relationship between 'Pumps' and 'AC System'?", "gt": "Yes"},
    {"id": 116, "q": "Which section comes immediately after 'Pipework'?", "gt": "Next Section"},
    {"id": 117, "q": "Are there more than 10 rules in Chapter 4?", "gt": "Yes/No"},
    {"id": 118, "q": "Does the structure link 'Valves' to 'Pipework' directly?", "gt": "Yes"},
    {"id": 119, "q": "Is 'Glossary' the first or last section?", "gt": "Last"},
    {"id": 120, "q": "Does the hierarchy support 'Sub-sub-sections' (e.g., 1.1.1)?", "gt": "Yes"},
    {"id": 121, "q": "Is 'Appendix A' a child of the root document node?", "gt": "Yes"},
    {"id": 122, "q": "Does the 'Definitions' section precede the 'Installation' section?", "gt": "Yes"},
    {"id": 123, "q": "Are 'Figures' listed as separate nodes in the hierarchy?", "gt": "No"},
    {"id": 124, "q": "Is there a 'Foreword' section in the 2022 code?", "gt": "Yes"},
    {"id": 125, "q": "Does the graph link 'British Standards' to specific chapters?", "gt": "Yes"},
    {"id": 126, "q": "Does the Fire Service 2012 code explicitly mention 'BS 8491' anywhere?", "gt": "No"},
    {"id": 127, "q": "Is there any rule in the 2020 Electrical Code that regulates 'Fiber Optic Cables'?", "gt": "No"},
    {"id": 128, "q": "Does the specification include material requirements for 'PVC pipes'?", "gt": "Yes/No"},
    {"id": 129, "q": "Are there any defined constraints for 'Ceiling fan' height in the 2012 edition?", "gt": "No"},
    {"id": 130, "q": "Does the code provide a testing method for 'Ductile iron' coating?", "gt": "Yes"},
    {"id": 131, "q": "Is 'Lead sheet' listed as a permitted material for acoustic lagging?", "gt": "Yes"},
    {"id": 132, "q": "Does the standard exclude 'Aluminium' for any specific application?", "gt": "Yes"},
    {"id": 133, "q": "Are there any temperature constraints defined for 'Saline water pumps'?", "gt": "Yes"},
    {"id": 134, "q": "Does the text mention 'Manufacturer A' as an approved vendor?", "gt": "No"},
    {"id": 135, "q": "Is 'Grade 316 Stainless Steel' explicitly required for any underwater installation?", "gt": "Yes"},
    {"id": 136, "q": "Does the code explicitly forbid 'Asbestos' materials?", "gt": "Yes"},
    {"id": 137, "q": "Is 'Copper' listed as an acceptable material for 'Fresh Water' pipes?", "gt": "Yes"},
    {"id": 138, "q": "Does the specification mention 'Solar Panels'?", "gt": "No"},
    {"id": 139, "q": "Are 'Gas Turbines' covered in the Mechanical spec?", "gt": "No"},
    {"id": 140, "q": "Is there a rule for 'Noise Control' in the Electrical section?", "gt": "No"},
    {"id": 141, "q": "Does the document list 'Approved Contractors'?", "gt": "No"},
    {"id": 142, "q": "Is 'Plastic' permitted for 'Fire Service' piping?", "gt": "No"},
    {"id": 143, "q": "Does the text define 'High Voltage' explicitly?", "gt": "Yes"},
    {"id": 144, "q": "Are '3D Printers' mentioned in the equipment list?", "gt": "No"},
    {"id": 145, "q": "Is 'WiFi' mentioned in the CCMS section?", "gt": "No"},
    {"id": 146, "q": "Does the code mention 'Smart Home' protocols?", "gt": "No"},
    {"id": 147, "q": "Is 'Bamboo' listed as a scaffolding material?", "gt": "No"},
    {"id": 148, "q": "Does the specification explicitly allow 'Recycled materials'?", "gt": "No"},
    {"id": 149, "q": "Is 'Titanium' mentioned for pipework?", "gt": "No"},
    {"id": 150, "q": "Are 'Nuclear' applications covered in this code?", "gt": "No"},
    {"id": 151, "q": "Which British Standards (BS) are referenced in the 'Cable' section?", "gt": "List"},
    {"id": 152, "q": "Does Rule R-FP-APP8-002 depend on any other constraints?", "gt": "Yes"},
    {"id": 153, "q": "List all rules that cite 'BS EN 60702'.", "gt": "List"},
    {"id": 154, "q": "Is the 'Fire Code' linked to the 'Water Supply' specification in the graph?", "gt": "No"},
    {"id": 155, "q": "Which other sections reference the 'General Specification'?", "gt": "List"},
    {"id": 156, "q": "Does the definition of 'Automatic actuating device' rely on an external standard?", "gt": "Yes"},
    {"id": 157, "q": "Are the standards for 'Emergency Generators' consistent between 2012 and 2022 codes?", "gt": "No"},
    {"id": 158, "q": "Which rule supersedes the 2012 requirement for cable fixings?", "gt": "2022 Rule"},
    {"id": 159, "q": "Do the 'Plumbing' and 'Fire Service' codes share any common material standards?", "gt": "Yes"},
    {"id": 160, "q": "Is the term 'Competent Person' defined in the Definitions chapter?", "gt": "Yes"},
    {"id": 161, "q": "Which BS standard is most frequently cited in the Mechanical code?", "gt": "BS Number"},
    {"id": 162, "q": "Does the 'Lighting' section reference the 'Energy Code'?", "gt": "Yes"},
    {"id": 163, "q": "Are 'Drawings' referenced as a requirement for approval?", "gt": "Yes"},
    {"id": 164, "q": "Does the 'Pump' section cross-reference 'Vibration Isolation'?", "gt": "Yes"},
    {"id": 165, "q": "Is 'BS 7671' referenced in the Fire Code?", "gt": "Yes"},
    {"id": 166, "q": "Do the 'Ventilation' rules reference 'Fire Dampers'?", "gt": "Yes"},
    {"id": 167, "q": "Is there a link between 'Switchboard' and 'Generator' sections?", "gt": "Yes"},
    {"id": 168, "q": "Does the 'Earthing' section reference 'Lightning Protection'?", "gt": "Yes"},
    {"id": 169, "q": "Are 'Safety' rules cross-referenced in the 'Installation' chapter?", "gt": "Yes"},
    {"id": 170, "q": "Does the 2022 code cite any 'ISO' standards?", "gt": "Yes"},
    {"id": 171, "q": "Is 'IEEE' referenced in the electrical section?", "gt": "Yes/No"},
    {"id": 172, "q": "Does the 'Alarm' section reference 'Telecommunications'?", "gt": "Yes"},
    {"id": 173, "q": "Are 'Buildings Dept' codes cited?", "gt": "Yes"},
    {"id": 174, "q": "Does the specification reference 'EMSD' guidelines?", "gt": "Yes"},
    {"id": 175, "q": "Is 'IEC 60364' cited alongside BS 7671?", "gt": "Yes"},
    {"id": 176, "q": "List all equipment that requires a 'minimum clearance' of 10mm or more.", "gt": "List"},
    {"id": 177, "q": "What is the maximum voltage rating mentioned in the Electrical Code?", "gt": "Value"},
    {"id": 178, "q": "List all pipe diameters that have specific 'joining method' rules.", "gt": "List"},
    {"id": 179, "q": "Which materials have a defined 'Density' property in the specifications?", "gt": "List"},
    {"id": 180, "q": "What are the different 'Classes' of Ductile Iron pipe mentioned?", "gt": "Classes"},
    {"id": 181, "q": "List all components that require 'Galvanised' coating.", "gt": "List"},
    {"id": 182, "q": "Which items in the 'Mechanical' section have a specified 'Swing Angle'?", "gt": "Hangers"},
    {"id": 183, "q": "What is the most common minimum height requirement in the dataset?", "gt": "Value"},
    {"id": 184, "q": "List all rules that involve a temperature threshold > 25°C.", "gt": "List"},
    {"id": 185, "q": "Which equipment types have 'vibration isolation' requirements?", "gt": "List"},
    {"id": 186, "q": "What is the maximum allowed noise level (dBA) mentioned?", "gt": "Value"},
    {"id": 187, "q": "How many rules specify a 'Safety Factor'?", "gt": "Count"},
    {"id": 188, "q": "List all items with a 'Service Life' requirement.", "gt": "List"},
    {"id": 189, "q": "Which cables have a 'Fire Survival Time' attribute?", "gt": "List"},
    {"id": 190, "q": "What is the smallest pipe size mentioned in the spec?", "gt": "Value"},
    {"id": 191, "q": "List all tests that require a 'Pressure' value.", "gt": "List"},
    {"id": 192, "q": "Which components have a 'Color Code' requirement?", "gt": "List"},
    {"id": 193, "q": "What is the maximum spacing for 'Hangers'?", "gt": "Value"},
    {"id": 194, "q": "List all materials that require 'Anti-corrosion' treatment.", "gt": "List"},
    {"id": 195, "q": "Which rules involve a 'Time' constraint (e.g., 60 mins)?", "gt": "List"},
    {"id": 196, "q": "Where is the requirement for 'Saddle' spacing defined? (Chapter/Section)", "gt": "Location"},
    {"id": 197, "q": "From which document does the rule regarding 'HV warning labels' originate?", "gt": "Doc Name"},
    {"id": 198, "q": "Is the rule for 'Pump foundations' found in the General Spec or the Particular Spec?", "gt": "General"},
    {"id": 199, "q": "Who is the issuing authority for the 'Fire Code' in the knowledge graph?", "gt": "FSD"},
    {"id": 200, "q": "Is 'Rule R-13-PIPE-001' active or deprecated?", "gt": "Active"},
    {"id": 201, "q": "According to Fire Code 2022, is BS 7629-1 applicable for a 19mm cable?", "gt": "Yes (<=20mm)"},
    {"id": 202, "q": "According to Fire Code 2022, is BS 7846 applicable for a 19mm cable?", "gt": "No (Only >20mm)"},
    {"id": 203, "q": "According to Fire Code 2022, is BS 7629-1 applicable for a 21mm cable?", "gt": "No (Only <=20mm)"},
    {"id": 204, "q": "According to Fire Code 2022, is BS 7846 applicable for a 21mm cable?", "gt": "Yes (>20mm)"},
    {"id": 205, "q": "Is 'Screwed joint' permitted for 48mm GS Fire pipe?", "gt": "Yes (<=50mm)"},
    {"id": 206, "q": "Is 'Screwed joint' permitted for 52mm GS Fire pipe?", "gt": "No (Use coupling)"},
    {"id": 207, "q": "Is 'Mechanical coupling' permitted for 48mm GS Fire pipe?", "gt": "No (Use screwed)"},
    {"id": 208, "q": "Is 'Mechanical coupling' permitted for 148mm GS Fire pipe?", "gt": "Yes (50-150mm)"},
    {"id": 209, "q": "Is 'Flanged joint' permitted for 148mm GS Fire pipe?", "gt": "No (Only >150mm)"},
    {"id": 210, "q": "Is 'Flanged joint' required for 152mm GS Fire pipe?", "gt": "Yes (>150mm)"},
    {"id": 211, "q": "Is steel pipe permitted for 79mm underground fire service?", "gt": "Yes (<80mm)"},
    {"id": 212, "q": "Is ductile iron required for 79mm underground fire service?", "gt": "No"},
    {"id": 213, "q": "Is steel pipe permitted for 81mm underground fire service?", "gt": "No (Must be DI)"},
    {"id": 214, "q": "Does BS 6387 Cat. CWZ require a 120-minute survival time in the text?", "gt": "No (Only specified for BS 8491)"},
    {"id": 215, "q": "Is BS 5839 mentioned as a cable standard in Appendix 8 Item 2?", "gt": "No"},
    {"id": 216, "q": "Can I use 'Copper' pipes for the main fire service underground ring?", "gt": "No (Steel/DI specified)"},
    {"id": 217, "q": "Is 'PVC' listed as an option for automatic actuating device cables?", "gt": "No"},
    {"id": 218, "q": "What is the category of BS 6387 required?", "gt": "Category CWZ"},
    {"id": 219, "q": "Is BS 6387 Category A acceptable?", "gt": "No (Must be CWZ)"},
    {"id": 220, "q": "Are 'Sprinkler waterflow switches' considered automatic actuating devices?", "gt": "Yes"},

    # --- Group 2: Electrical Code (Precision Numeric Checks) ---
    {"id": 221, "q": "Is a 29mm HV label compliant in 2015?", "gt": "No (Min 30mm)"},
    {"id": 222, "q": "Is a 31mm HV label compliant in 2015?", "gt": "Yes"},
    {"id": 223, "q": "Is a 49mm HV label compliant in 2020?", "gt": "No (Min 50mm)"},
    {"id": 224, "q": "Is a 51mm HV label compliant in 2020?", "gt": "Yes"},
    {"id": 225, "q": "Can saddle spacing be exactly 150mm without intermediate fixing?", "gt": "Yes (<=150mm)"},
    {"id": 226, "q": "Can saddle spacing be 151mm without intermediate fixing?", "gt": "No"},
    {"id": 227, "q": "Is a screw protrusion of exactly 3 threads acceptable?", "gt": "Yes"},
    {"id": 228, "q": "Is a screw protrusion of 3.5 threads acceptable?", "gt": "No"},
    {"id": 229, "q": "Is 1.2m support spacing allowed for general tray?", "gt": "Yes (<=1.2m)"},
    {"id": 230, "q": "Is 1.25m support spacing allowed for general tray?", "gt": "No"},
    {"id": 231, "q": "Is 1.5m support spacing allowed for wire mesh tray?", "gt": "Yes (<=1.5m)"},
    {"id": 232, "q": "Is 1.55m support spacing allowed for wire mesh tray?", "gt": "No"},
    {"id": 233, "q": "Is 9mm clearance acceptable for cable loops?", "gt": "No (Min 10mm)"},
    {"id": 234, "q": "Is 11mm clearance acceptable for cable loops?", "gt": "Yes"},
    {"id": 235, "q": "Is 19mm conduit allowed?", "gt": "No (Min 20mm)"},
    {"id": 236, "q": "Is 21mm conduit allowed?", "gt": "Yes"},
    {"id": 237, "q": "Does a 290mm tray run require 1 support?", "gt": "Yes"},
    {"id": 238, "q": "Does a 290mm tray run require 2 supports?", "gt": "No (Short run)"},
    {"id": 239, "q": "Does a 310mm tray run require 1 support?", "gt": "No (Min 2)"},
    {"id": 240, "q": "Does a 310mm tray run require 2 supports?", "gt": "Yes"},

    # --- Group 3: Mechanical (Fan & CCMS Precision) ---
    {"id": 241, "q": "Is 2390mm height allowed for ceiling fan?", "gt": "No (Min 2400mm)"},
    {"id": 242, "q": "Is 2410mm height allowed for ceiling fan?", "gt": "Yes"},
    {"id": 243, "q": "Is 2590mm height allowed for ceiling fan?", "gt": "Yes"},
    {"id": 244, "q": "Is 2610mm height allowed for ceiling fan?", "gt": "No (Max 2600mm)"},
    {"id": 245, "q": "Can CCMS off time be 4%?", "gt": "No (Min 5%)"},
    {"id": 246, "q": "Can CCMS off time be 6%?", "gt": "Yes"},
    {"id": 247, "q": "Can CCMS off time be 94%?", "gt": "Yes"},
    {"id": 248, "q": "Can CCMS off time be 96%?", "gt": "No"},
    {"id": 249, "q": "Is 9mm clearance compliant for spring mount?", "gt": "No (Min 10mm)"},
    {"id": 250, "q": "Is 11mm clearance compliant for spring mount?", "gt": "Yes"},
    {"id": 251, "q": "Is 29 degrees swing angle valid for spring hanger?", "gt": "No (30-35)"},
    {"id": 252, "q": "Is 31 degrees swing angle valid for spring hanger?", "gt": "Yes"},
    {"id": 253, "q": "Is 34 degrees swing angle valid for spring hanger?", "gt": "Yes"},
    {"id": 254, "q": "Is 36 degrees swing angle valid for spring hanger?", "gt": "No"},
    {"id": 255, "q": "Is 24mm seal projection compliant?", "gt": "No (Min 25mm)"},
    {"id": 256, "q": "Is 26mm seal projection compliant?", "gt": "Yes"},
    {"id": 257, "q": "Is flared joint allowed for 20mm copper pipe?", "gt": "Yes (<=20mm)"},
    {"id": 258, "q": "Is flared joint allowed for 22mm copper pipe?", "gt": "No"},
    {"id": 259, "q": "Is 590mm distance for local control box acceptable?", "gt": "Yes (<=600mm)"},
    {"id": 260, "q": "Is 610mm distance for local control box acceptable?", "gt": "No"},

    # --- Group 4: Plumbing & Drainage (Pump Ratios & Coatings) ---
    {"id": 261, "q": "Is 2.4x pump mass ratio acceptable in Plumbing?", "gt": "No (Min 2.5x)"},
    {"id": 262, "q": "Is 2.5x pump mass ratio acceptable in Plumbing?", "gt": "Yes"},
    {"id": 263, "q": "Is 90mm inertia block thickness acceptable?", "gt": "No (Min 100mm)"},
    {"id": 264, "q": "Is 110mm inertia block thickness acceptable?", "gt": "Yes"},
    {"id": 265, "q": "Is 140mm inertia block projection acceptable?", "gt": "No (Min 150mm)"},
    {"id": 266, "q": "Is 160mm inertia block projection acceptable?", "gt": "Yes"},
    {"id": 267, "q": "Is 40mm base clearance acceptable?", "gt": "No (Min 50mm)"},
    {"id": 268, "q": "Is 60mm base clearance acceptable?", "gt": "No (Must be 50mm)"},
    {"id": 269, "q": "Is 'HAC mortar' allowed for Plumbing DI pipe internal?", "gt": "No (Cement lining)"},
    {"id": 270, "q": "Is 'Cement lining' allowed for Drainage DI pipe internal?", "gt": "No (HAC mortar)"},
    {"id": 271, "q": "Is 'Bitumen' allowed for Fire DI pipe external?", "gt": "No (Zinc)"},
    {"id": 272, "q": "Is 'Zinc' allowed for Plumbing DI pipe external?", "gt": "No (Bitumen)"},
    {"id": 273, "q": "Is 0.45x rated deflection reserve acceptable?", "gt": "No (Min 0.5x)"},
    {"id": 274, "q": "Is 0.55x rated deflection reserve acceptable?", "gt": "Yes"},
    {"id": 275, "q": "Is 'Screwed joint' permitted for 90mm Plumbing GS pipe?", "gt": "Yes (<=100mm)"},
    {"id": 276, "q": "Is 'Screwed joint' permitted for 110mm Plumbing GS pipe?", "gt": "No"},
    {"id": 277, "q": "Is 'Flanged joint' required for 140mm Plumbing GS pipe?", "gt": "No (Only >=150mm)"},
    {"id": 278, "q": "Is 'Flanged joint' required for 160mm Plumbing GS pipe?", "gt": "Yes"},
    {"id": 279, "q": "Is 'Oversized sleeve' required for 140mm sleeve?", "gt": "No (Standard)"},
    {"id": 280, "q": "Is 'Oversized sleeve' required for 160mm sleeve?", "gt": "Yes"},

    # --- Group 5: Electrical Systems (Transformers & Cables) ---
    {"id": 281, "q": "Is 2.4x starting current allowed for auto-transformers?", "gt": "Yes (Max 2.5x)"},
    {"id": 282, "q": "Is 2.6x starting current allowed for auto-transformers?", "gt": "No"},
    {"id": 283, "q": "Is 45kV impulse voltage specified?", "gt": "Yes"},
    {"id": 284, "q": "Is 50kV impulse voltage specified?", "gt": "No (45kV)"},
    {"id": 285, "q": "Is 60% tapping supported?", "gt": "Yes"},
    {"id": 286, "q": "Is 65% tapping supported?", "gt": "No (60, 75, 85)"},
    {"id": 287, "q": "Is 70% tapping supported?", "gt": "No"},
    {"id": 288, "q": "Is 75% tapping supported?", "gt": "Yes"},
    {"id": 289, "q": "Is 80% tapping supported?", "gt": "No"},
    {"id": 290, "q": "Is 85% tapping supported?", "gt": "Yes"},
    {"id": 291, "q": "Is 'Indoor' type required for auto-transformers?", "gt": "Yes"},
    {"id": 292, "q": "Is 'Outdoor' type allowed for auto-transformers?", "gt": "No"},
    {"id": 293, "q": "Is 'Natural air cooled' required?", "gt": "Yes"},
    {"id": 294, "q": "Is 'Forced air cooled' required?", "gt": "No"},
    {"id": 295, "q": "Is 3.3kV system voltage allowed?", "gt": "Yes"},
    {"id": 296, "q": "Is 6.6kV system voltage allowed?", "gt": "Yes"},
    {"id": 297, "q": "Is 11kV system voltage allowed?", "gt": "Yes"},
    {"id": 298, "q": "Is 22kV system voltage allowed?", "gt": "No (Not listed)"},
    {"id": 299, "q": "Is 450V cable grade acceptable for 3-phase?", "gt": "Yes"},
    {"id": 300, "q": "Is 750V cable grade acceptable for 3-phase?", "gt": "Yes"},

    # --- Group 6: Logic Trap Questions (Cross-checking) ---
    {"id": 301, "q": "Does the Plumbing code apply to 'Fire Sprinkler' pipes?", "gt": "No (Fire Code)"},
    {"id": 302, "q": "Does the Fire Code apply to 'Fresh Water' pipes?", "gt": "No (Plumbing Code)"},
    {"id": 303, "q": "Is 'HAC mortar' used for Plumbing pipes?", "gt": "No"},
    {"id": 304, "q": "Is 'Cement lining' used for Drainage pipes?", "gt": "No"},
    {"id": 305, "q": "Is 'Zinc' coating used for Plumbing pipes?", "gt": "No"},
    {"id": 306, "q": "Is 'Bitumen' coating used for Fire pipes?", "gt": "No"},
    {"id": 307, "q": "Do Electrical regulations specify 'Pump Foundation' mass?", "gt": "No"},
    {"id": 308, "q": "Do Mechanical regulations specify 'HV Label' size?", "gt": "No"},
    {"id": 309, "q": "Is 'Galvanised Steel' permitted for underground Fire service >=80mm?", "gt": "No (Ductile Iron)"},
    {"id": 310, "q": "Is 'Steel' permitted for underground Fire service <80mm?", "gt": "Yes"},
    {"id": 311, "q": "Is 120g/m2 zinc coating compliant for Fire DI pipe?", "gt": "No (Min 130g)"},
    {"id": 312, "q": "Is 130g/m2 zinc coating compliant for Fire DI pipe?", "gt": "Yes"},
    {"id": 313, "q": "Is 99.8% zinc purity compliant?", "gt": "No (Min 99.9%)"},
    {"id": 314, "q": "Is 99.9% zinc purity compliant?", "gt": "Yes"},
    {"id": 315, "q": "Does the 2012 Fire code mention 'BS 8491'?", "gt": "No"},
    {"id": 316, "q": "Does the 2022 Fire code mention 'BS 8491'?", "gt": "Yes"},
    {"id": 317, "q": "Does 2015 Electrical Code require 50mm HV labels?", "gt": "No (30mm)"},
    {"id": 318, "q": "Does 2020 Electrical Code require 30mm HV labels?", "gt": "No (50mm)"},
    {"id": 319, "q": "Is 'Lead sheet' required for 'Thermal insulation'?", "gt": "No (Acoustic)"},
    {"id": 320, "q": "Is 'Aluminium jacket' required for 'Acoustic duct lag'?", "gt": "Yes"},

    # --- Group 7: Material & Specification Validity ---
    {"id": 321, "q": "Is 'Hot-dip galvanised' steel allowed for tray supports?", "gt": "Yes"},
    {"id": 322, "q": "Is 'Black steel' allowed for tray supports?", "gt": "No (Must be treated)"},
    {"id": 323, "q": "Is 'Anti-rust epoxy' allowed for tray supports?", "gt": "Yes"},
    {"id": 324, "q": "Is 'Anti-rusting metal' allowed for tray supports?", "gt": "Yes"},
    {"id": 325, "q": "Is 'Copper' allowed for trunking wiring?", "gt": "Yes"},
    {"id": 326, "q": "Is 'Aluminium' allowed for trunking wiring?", "gt": "No"},
    {"id": 327, "q": "Is 'Glass fibre' density 20 kg/m3 acceptable?", "gt": "No (24 kg/m3)"},
    {"id": 328, "q": "Is 'Glass fibre' density 24 kg/m3 acceptable?", "gt": "Yes"},
    {"id": 329, "q": "Is 'Lead sheet' weight 4 kg/m2 acceptable?", "gt": "No (5 kg/m2)"},
    {"id": 330, "q": "Is 'Lead sheet' weight 5 kg/m2 acceptable?", "gt": "Yes"},
    {"id": 331, "q": "Is 'Barium-loaded vinyl' allowed?", "gt": "Yes"},
    {"id": 332, "q": "Is 'Stainless steel' required for fresh water?", "gt": "No"},
    {"id": 333, "q": "Is 'Stainless steel' required for saline water?", "gt": "Yes (if >28C)"},
    {"id": 334, "q": "Is 'Grade 304' stainless steel specified?", "gt": "No (Usually 316 for saline)"},
    {"id": 335, "q": "Is 'IEC 62271-106' the standard for transformers?", "gt": "Yes"},
    {"id": 336, "q": "Is 'BS 7671' the standard for transformers?", "gt": "No"},
    {"id": 337, "q": "Is 'BS 6387' the standard for fire cables?", "gt": "Yes"},
    {"id": 338, "q": "Is 'BS 5839' the standard for fire cables in this text?", "gt": "No"},
    {"id": 339, "q": "Is 'BS 1363' mentioned for sockets?", "gt": "No"},
    {"id": 340, "q": "Is 'BS EN 60702' mentioned for cables?", "gt": "Yes"},

    # --- Group 8: Application Context ---
    {"id": 341, "q": "Is a loop required for cable entry to 'Vibrating equipment'?", "gt": "Yes"},
    {"id": 342, "q": "Is a loop required for cable entry to 'Static equipment'?", "gt": "No"},
    {"id": 343, "q": "Is 2.5x mass ratio for 'Sump pump'?", "gt": "Yes"},
    {"id": 344, "q": "Is 2.5x mass ratio for 'Chiller'?", "gt": "No"},
    {"id": 345, "q": "Is 30-35 degree swing for 'Spring Hanger'?", "gt": "Yes"},
    {"id": 346, "q": "Is 30-35 degree swing for 'Spring Mount'?", "gt": "No"},
    {"id": 347, "q": "Is 10mm clearance for 'Spring Mount'?", "gt": "Yes"},
    {"id": 348, "q": "Is 10mm clearance for 'Spring Hanger'?", "gt": "No"},
    {"id": 349, "q": "Is 50mm clearance for 'Pump Base'?", "gt": "Yes"},
    {"id": 350, "q": "Is 50mm clearance for 'Cable Tray'?", "gt": "No (20mm)"},
    {"id": 351, "q": "Is 20mm clearance for 'Cable Tray'?", "gt": "Yes"},
    {"id": 352, "q": "Is 20mm clearance for 'Pump Base'?", "gt": "No (50mm)"},
    {"id": 353, "q": "Is 25mm projection for 'Wall Seals'?", "gt": "Yes"},
    {"id": 354, "q": "Is 25mm projection for 'Pump Inertia'?", "gt": "No (150mm)"},
    {"id": 355, "q": "Is 150mm projection for 'Pump Inertia'?", "gt": "Yes"},
    {"id": 356, "q": "Is 100mm thickness for 'Pump Inertia'?", "gt": "Yes"},
    {"id": 357, "q": "Is 50mm thickness for 'Acoustic Lag'?", "gt": "Yes"},
    {"id": 358, "q": "Is 50mm thickness for 'Pump Inertia'?", "gt": "No (100mm)"},
    {"id": 359, "q": "Is 1.2m spacing for 'General Tray'?", "gt": "Yes"},
    {"id": 360, "q": "Is 1.5m spacing for 'Wire Mesh Tray'?", "gt": "Yes"},

    # --- Group 9: Document & Version Checks ---
    {"id": 361, "q": "Is 'J001' from Fire Code 2022?", "gt": "Yes"},
    {"id": 362, "q": "Is 'J002' from Fire Code 2012?", "gt": "Yes"},
    {"id": 363, "q": "Is 'J003' from Electrical Code 2020?", "gt": "Yes"},
    {"id": 364, "q": "Is 'J004' from Electrical Code 2015?", "gt": "Yes"},
    {"id": 365, "q": "Is 'J007' from Drainage Spec?", "gt": "Yes"},
    {"id": 366, "q": "Is 'J008' from Plumbing Spec?", "gt": "Yes"},
    {"id": 367, "q": "Is 'J014' from Mechanical Spec?", "gt": "Yes"},
    {"id": 368, "q": "Is 'J018' from Electrical Spec?", "gt": "Yes"},
    {"id": 369, "q": "Is 'J020' from Mechanical Spec?", "gt": "Yes"},
    {"id": 370, "q": "Is 'J026' from Electrical Spec?", "gt": "Yes"},
    {"id": 371, "q": "Does J001 replace J002 rules?", "gt": "Yes (2022 vs 2012)"},
    {"id": 372, "q": "Does J003 replace J004 rules?", "gt": "Yes (2020 vs 2015)"},
    {"id": 373, "q": "Does J005 replace J006 rules?", "gt": "Yes (2020 vs 2015)"},
    {"id": 374, "q": "Are J007 and J008 from the same document?", "gt": "No (Drainage vs Plumbing)"},
    {"id": 375, "q": "Are J011 and J012 from the same document?", "gt": "No (Drainage vs Plumbing)"},
    {"id": 376, "q": "Are J014 and J015 from the same document?", "gt": "Yes (Mechanical)"},
    {"id": 377, "q": "Are J022 and J023 from the same document?", "gt": "Yes (CCMS)"},
    {"id": 378, "q": "Are J028 and J030 from the same document?", "gt": "Yes (Cable Trays)"},
    {"id": 379, "q": "Is J010 from Fire Service Spec?", "gt": "Yes"},
    {"id": 380, "q": "Is J009 from Plumbing Spec?", "gt": "Yes"},

    # --- Group 10: Miscellaneous Logic ---
    {"id": 381, "q": "Does 'Wiring' fall under Mechanical?", "gt": "No"},
    {"id": 382, "q": "Does 'Pumps' fall under Electrical?", "gt": "No (Mechanical/Plumbing)"},
    {"id": 383, "q": "Does 'CCMS' fall under Mechanical?", "gt": "Yes"},
    {"id": 384, "q": "Does 'Acoustic Lag' fall under Mechanical?", "gt": "Yes"},
    {"id": 385, "q": "Does 'HV Switchgear' fall under Electrical?", "gt": "Yes"},
    {"id": 386, "q": "Does 'Trunking' fall under Electrical?", "gt": "Yes"},
    {"id": 387, "q": "Does 'Fire Pipe' fall under Fire Services?", "gt": "Yes"},
    {"id": 388, "q": "Does 'Ductile Iron' fall under Material?", "gt": "Yes"},
    {"id": 389, "q": "Is 'BS 7846' a British Standard?", "gt": "Yes"},
    {"id": 390, "q": "Is 'IEC' an International Standard?", "gt": "Yes"},
    {"id": 391, "q": "Is '2022' the latest Fire Code mentioned?", "gt": "Yes"},
    {"id": 392, "q": "Is '2020' the latest Electrical Code mentioned?", "gt": "Yes"},
    {"id": 393, "q": "Is '2017' the latest Plumbing Code mentioned?", "gt": "Yes"},
    {"id": 394, "q": "Is '2028' a valid edition for Cable Trays?", "gt": "Yes (As per J028 text)"},
    {"id": 395, "q": "Is '2012' an obsolete Fire Code?", "gt": "Yes"},
    {"id": 396, "q": "Is '2015' an obsolete Electrical Code?", "gt": "Yes"},
    {"id": 397, "q": "Is 'Appendix 8' relevant to Cables?", "gt": "Yes"},
    {"id": 398, "q": "Is 'Part B3' relevant to Drainage?", "gt": "Yes"},
    {"id": 399, "q": "Is 'Part B2.5' relevant to Plumbing?", "gt": "Yes"},
    {"id": 400, "q": "Is 'Section 6' relevant to Mechanical?", "gt": "Yes"},
    {"id": 401, "q": "Is 'Auto-star' connection mandatory for auto-transformers?", "gt": "Yes"},
    {"id": 402, "q": "Is 'Delta-delta' connection permitted for auto-transformers?", "gt": "No (Auto-star specified)"},
    {"id": 403, "q": "Is a starting current of 2.0x full load compliant?", "gt": "Yes (<=2.5x)"},
    {"id": 404, "q": "Is a starting current of 2.5x full load compliant?", "gt": "Yes"},
    {"id": 405, "q": "Is a starting current of 2.6x full load compliant?", "gt": "No (Max 2.5x)"},
    {"id": 406, "q": "Does the spec allow 60% tapping?", "gt": "Yes"},
    {"id": 407, "q": "Does the spec allow 65% tapping?", "gt": "No"},
    {"id": 408, "q": "Does the spec allow 70% tapping?", "gt": "No"},
    {"id": 409, "q": "Does the spec allow 75% tapping?", "gt": "Yes"},
    {"id": 410, "q": "Does the spec allow 85% tapping?", "gt": "Yes"},
    {"id": 411, "q": "Is 'Natural air cooled' the required cooling method?", "gt": "Yes"},
    {"id": 412, "q": "Is 'Force air cooled' an acceptable alternative?", "gt": "No"},
    {"id": 413, "q": "Is 45kV peak impulse voltage required?", "gt": "Yes"},
    {"id": 414, "q": "Is 30kV peak impulse voltage sufficient?", "gt": "No"},
    {"id": 415, "q": "Is the impulse duration 1/50 microseconds?", "gt": "Yes"},
    {"id": 416, "q": "Is 50Hz the operating frequency?", "gt": "Yes"},
    {"id": 417, "q": "Is 60Hz the operating frequency?", "gt": "No"},
    {"id": 418, "q": "Are these transformers for 'Squirrel cage' motors?", "gt": "Yes"},
    {"id": 419, "q": "Are these transformers for 'Slip ring' motors?", "gt": "No"},
    {"id": 420, "q": "Is 'No breaking' starting required?", "gt": "Yes"},

    # --- Group 2: Cable Tray & Trunking (Support Logic) ---
    {"id": 421, "q": "Does a 200mm cable tray run require 1 support?", "gt": "Yes (Run <300mm)"},
    {"id": 422, "q": "Does a 200mm cable tray run require 2 supports?", "gt": "No (1 is sufficient)"},
    {"id": 423, "q": "Does a 400mm cable tray run require 1 support?", "gt": "No (Run >=300mm needs 2)"},
    {"id": 424, "q": "Does a 400mm cable tray run require 2 supports?", "gt": "Yes"},
    {"id": 425, "q": "Is 1.2m spacing allowed for general tray?", "gt": "Yes"},
    {"id": 426, "q": "Is 1.3m spacing allowed for general tray?", "gt": "No"},
    {"id": 427, "q": "Is 1.5m spacing allowed for wire mesh tray?", "gt": "Yes"},
    {"id": 428, "q": "Is 1.6m spacing allowed for wire mesh tray?", "gt": "No"},
    {"id": 429, "q": "Is 20mm clearance required behind trays?", "gt": "Yes"},
    {"id": 430, "q": "Is 15mm clearance required behind trays?", "gt": "No (Min 20mm)"},
    {"id": 431, "q": "Is 'Anti-rusting metal' allowed for brackets?", "gt": "Yes"},
    {"id": 432, "q": "Is 'Painted wood' allowed for brackets?", "gt": "No"},
    {"id": 433, "q": "Is 'Hot-dip galvanised' allowed for brackets?", "gt": "Yes"},
    {"id": 434, "q": "Is 450/750V cable grade required for 3-phase trunking?", "gt": "Yes"},
    {"id": 435, "q": "Is 300/500V cable grade allowed for 3-phase trunking?", "gt": "No"},
    {"id": 436, "q": "Are non-sheathed copper cables allowed in trunking?", "gt": "Yes"},
    {"id": 437, "q": "Are aluminium cables allowed in trunking?", "gt": "No"},
    {"id": 438, "q": "Is fixing distance 225mm from bends?", "gt": "Yes"},
    {"id": 439, "q": "Is fixing distance 300mm from bends?", "gt": "No (Max 225mm)"},
    {"id": 440, "q": "Is 'Epoxy coating' allowed for tray supports?", "gt": "Yes"},

    # --- Group 3: Mechanical (CCMS & Vibration Precision) ---
    {"id": 441, "q": "Is 5% off-time valid for CCMS?", "gt": "Yes"},
    {"id": 442, "q": "Is 4% off-time valid for CCMS?", "gt": "No (Min 5%)"},
    {"id": 443, "q": "Is 95% off-time valid for CCMS?", "gt": "Yes"},
    {"id": 444, "q": "Is 96% off-time valid for CCMS?", "gt": "No (Max 95%)"},
    {"id": 445, "q": "Is 1 minute time resolution valid?", "gt": "Yes"},
    {"id": 446, "q": "Is 120 minutes time resolution valid?", "gt": "Yes"},
    {"id": 447, "q": "Is 121 minutes time resolution valid?", "gt": "No (Max 120)"},
    {"id": 448, "q": "Is 30 degree swing angle valid?", "gt": "Yes"},
    {"id": 449, "q": "Is 29 degree swing angle valid?", "gt": "No (Min 30)"},
    {"id": 450, "q": "Is 35 degree swing angle valid?", "gt": "Yes"},
    {"id": 451, "q": "Is 36 degree swing angle valid?", "gt": "No (Max 35)"},
    {"id": 452, "q": "Is 10mm clearance for bolts required?", "gt": "Yes"},
    {"id": 453, "q": "Is 5mm clearance for bolts sufficient?", "gt": "No"},
    {"id": 454, "q": "Is 'Stainless steel' required for fresh water pumps?", "gt": "No"},
    {"id": 455, "q": "Is 'Stainless steel' required for saline water >28C?", "gt": "Yes"},
    {"id": 456, "q": "Is 24kg/m3 density for fibreglass?", "gt": "Yes"},
    {"id": 457, "q": "Is 20kg/m3 density for fibreglass acceptable?", "gt": "No"},
    {"id": 458, "q": "Is 0.032 W/mC thermal conductivity max?", "gt": "Yes"},
    {"id": 459, "q": "Is 0.035 W/mC thermal conductivity acceptable?", "gt": "No"},
    {"id": 460, "q": "Is 'Factory-applied aluminium jacket' required?", "gt": "Yes"},

    # --- Group 4: Plumbing & Drainage (Pipe & Pump) ---
    {"id": 461, "q": "Is 'HAC mortar' allowed for Fire pipe internal?", "gt": "No (Drainage only)"},
    {"id": 462, "q": "Is 'Cement lining' allowed for Drainage pipe internal?", "gt": "No (Plumbing only)"},
    {"id": 463, "q": "Is 'Zinc' allowed for Drainage pipe external?", "gt": "Yes"},
    {"id": 464, "q": "Is 'Bitumen' allowed for Fire pipe external?", "gt": "No (Zinc)"},
    {"id": 465, "q": "Is 2.5x mass ratio required for Drainage pumps?", "gt": "No (Plumbing only)"},
    {"id": 466, "q": "Is 2.5x mass ratio required for Plumbing pumps?", "gt": "Yes"},
    {"id": 467, "q": "Is 100mm inertia block thickness for Plumbing?", "gt": "Yes"},
    {"id": 468, "q": "Is 150mm inertia block projection for Plumbing?", "gt": "Yes"},
    {"id": 469, "q": "Is 50mm base clearance for Plumbing?", "gt": "Yes"},
    {"id": 470, "q": "Is 0.5x deflection reserve for Plumbing?", "gt": "Yes"},
    {"id": 471, "q": "Is screwed joint allowed for 90mm Plumbing GS pipe?", "gt": "Yes (<=100mm)"},
    {"id": 472, "q": "Is screwed joint allowed for 90mm Fire GS pipe?", "gt": "No (<=50mm)"},
    {"id": 473, "q": "Is flanged joint allowed for 160mm Plumbing GS pipe?", "gt": "Yes (>=150mm)"},
    {"id": 474, "q": "Is flanged joint allowed for 160mm Fire GS pipe?", "gt": "Yes (>150mm)"},
    {"id": 475, "q": "Is standard sleeve allowed for 140mm pipe?", "gt": "Yes"},
    {"id": 476, "q": "Is oversized sleeve allowed for 160mm pipe?", "gt": "Yes"},
    {"id": 477, "q": "Is 130g/m2 zinc coating required for Fire DI pipe?", "gt": "Yes"},
    {"id": 478, "q": "Is 99.9% zinc purity required for Fire DI pipe?", "gt": "Yes"},
    {"id": 479, "q": "Is steel pipe allowed for 50mm underground fire service?", "gt": "Yes"},
    {"id": 480, "q": "Is DI pipe required for 100mm underground fire service?", "gt": "Yes"},

    # --- Group 5: Electrical (Labels & Saddles) ---
    {"id": 481, "q": "Is 30mm HV label allowed in 2015?", "gt": "Yes"},
    {"id": 482, "q": "Is 29mm HV label allowed in 2015?", "gt": "No"},
    {"id": 483, "q": "Is 50mm HV label allowed in 2020?", "gt": "Yes"},
    {"id": 484, "q": "Is 49mm HV label allowed in 2020?", "gt": "No"},
    {"id": 485, "q": "Is 3m label spacing allowed?", "gt": "Yes"},
    {"id": 486, "q": "Is 3.1m label spacing allowed?", "gt": "No"},
    {"id": 487, "q": "Is 150mm saddle spacing allowed?", "gt": "Yes"},
    {"id": 488, "q": "Is 151mm saddle spacing allowed without fixings?", "gt": "No"},
    {"id": 489, "q": "Is 3 threads protrusion allowed?", "gt": "Yes"},
    {"id": 490, "q": "Is 4 threads protrusion allowed?", "gt": "No"},
    {"id": 491, "q": "Is 20mm conduit diameter allowed?", "gt": "Yes"},
    {"id": 492, "q": "Is 19mm conduit diameter allowed?", "gt": "No"},
    {"id": 493, "q": "Is 60W load per lampholder allowed in 2020?", "gt": "Yes"},
    {"id": 494, "q": "Is 100W load per lampholder required in 2015?", "gt": "Yes"},
    {"id": 495, "q": "Is 10mm loop clearance required?", "gt": "Yes"},
    {"id": 496, "q": "Is 5mm loop clearance acceptable?", "gt": "No"},
    {"id": 497, "q": "Is 'Cable loop' for static equipment?", "gt": "No (Vibrating only)"},
    {"id": 498, "q": "Is 'Cable loop' for vibrating equipment?", "gt": "Yes"},
    {"id": 499, "q": "Is IEC 62271-106 for Switchgear?", "gt": "Yes"},
    {"id": 500, "q": "Is BS 7671 for Switchgear?", "gt": "No"},

    # --- Group 6: Fire Service (Cable Standards Logic) ---
    {"id": 501, "q": "Is BS 7846 for cables >20mm?", "gt": "Yes"},
    {"id": 502, "q": "Is BS 7846 for cables <=20mm?", "gt": "No"},
    {"id": 503, "q": "Is BS 7629-1 for cables <=20mm?", "gt": "Yes"},
    {"id": 504, "q": "Is BS 7629-1 for cables >20mm?", "gt": "No"},
    {"id": 505, "q": "Is BS 6387 Cat CWZ allowed?", "gt": "Yes"},
    {"id": 506, "q": "Is BS 6387 Cat A allowed?", "gt": "No"},
    {"id": 507, "q": "Is BS 8491 survival time 120 mins?", "gt": "Yes"},
    {"id": 508, "q": "Is BS 8491 survival time 60 mins?", "gt": "No"},
    {"id": 509, "q": "Does 2022 Fire Code include BS 6207?", "gt": "No"},
    {"id": 510, "q": "Does 2012 Fire Code include BS 6207?", "gt": "Yes"},
    {"id": 511, "q": "Is BS EN 60702 allowed in 2022?", "gt": "Yes"},
    {"id": 512, "q": "Is 'Automatic actuating device' cable covered?", "gt": "Yes"},
    {"id": 513, "q": "Is 'Sprinkler switch' an actuating device?", "gt": "Yes"},
    {"id": 514, "q": "Is 'Heat detector' an actuating device?", "gt": "Yes"},
    {"id": 515, "q": "Is 'Smoke detector' an actuating device?", "gt": "Yes"},
    {"id": 516, "q": "Is 'Emergency Generator' cable covered in 2012?", "gt": "Yes"},
    {"id": 517, "q": "Is 'PH rating' required in 2012?", "gt": "No"},
    {"id": 518, "q": "Is '20mm' the diameter threshold?", "gt": "Yes"},
    {"id": 519, "q": "Is '25mm' the diameter threshold?", "gt": "No"},
    {"id": 520, "q": "Is 'Mechanical coupling' allowed for 100mm Fire pipe?", "gt": "Yes"},

    # --- Group 7: Material Hallucination Checks (Context) ---
    {"id": 521, "q": "Is 'Lead sheet' specified for Fire pipes?", "gt": "No"},
    {"id": 522, "q": "Is 'Lead sheet' specified for Acoustic lag?", "gt": "Yes"},
    {"id": 523, "q": "Is 'HAC mortar' specified for Electrical conduit?", "gt": "No"},
    {"id": 524, "q": "Is 'HAC mortar' specified for Drainage?", "gt": "Yes"},
    {"id": 525, "q": "Is 'Bitumen' specified for Ceiling fans?", "gt": "No"},
    {"id": 526, "q": "Is 'Bitumen' specified for Plumbing?", "gt": "Yes"},
    {"id": 527, "q": "Is 'Zinc' specified for Auto-transformers?", "gt": "No"},
    {"id": 528, "q": "Is 'Zinc' specified for Fire pipes?", "gt": "Yes"},
    {"id": 529, "q": "Is 'Stainless steel' specified for Tray?", "gt": "No (HDG/Epoxy)"},
    {"id": 530, "q": "Is 'Stainless steel' specified for Saline Pump?", "gt": "Yes"},
    {"id": 531, "q": "Is 'Barium vinyl' specified for Drainage?", "gt": "No"},
    {"id": 532, "q": "Is 'Barium vinyl' specified for Acoustic lag?", "gt": "Yes"},
    {"id": 533, "q": "Is 'Glass fibre' specified for Pumps?", "gt": "No"},
    {"id": 534, "q": "Is 'Glass fibre' specified for Acoustic lag?", "gt": "Yes"},
    {"id": 535, "q": "Is 'Copper' specified for Trunking?", "gt": "Yes"},
    {"id": 536, "q": "Is 'Copper' specified for Fire Main?", "gt": "No (GS/DI)"},
    {"id": 537, "q": "Is 'Steel' specified for Conduit?", "gt": "Yes"},
    {"id": 538, "q": "Is 'Plastic' specified for Fire Conduit?", "gt": "No (Not in text)"},
    {"id": 539, "q": "Is 'Flanged joint' specified for Copper pipe?", "gt": "Yes"},
    {"id": 540, "q": "Is 'Welded joint' specified for Copper pipe?", "gt": "No (Brazed)"},

    # --- Group 8: Document Structure & Meta ---
    {"id": 541, "q": "Is 'J001' a Fire Code chapter?", "gt": "Yes"},
    {"id": 542, "q": "Is 'J002' a Fire Code chapter?", "gt": "Yes"},
    {"id": 543, "q": "Is 'J003' an Electrical Code chapter?", "gt": "Yes"},
    {"id": 544, "q": "Is 'J007' a Drainage Code chapter?", "gt": "Yes"},
    {"id": 545, "q": "Is 'J008' a Plumbing Code chapter?", "gt": "Yes"},
    {"id": 546, "q": "Is 'J014' a Mechanical Code chapter?", "gt": "Yes"},
    {"id": 547, "q": "Is 'J018' an Electrical Code chapter?", "gt": "Yes"},
    {"id": 548, "q": "Is 'Appendix 8' in Fire Code?", "gt": "Yes"},
    {"id": 549, "q": "Is 'Code 17A' in Electrical Code?", "gt": "Yes"},
    {"id": 550, "q": "Is 'Part B3' in Drainage Code?", "gt": "Yes"},
    {"id": 551, "q": "Is 'Part B2.5' in Plumbing Code?", "gt": "Yes"},
    {"id": 552, "q": "Is 'Section 6' in Mechanical Code?", "gt": "Yes"},
    {"id": 553, "q": "Is 'Section 7' in Electrical Code?", "gt": "Yes"},
    {"id": 554, "q": "Does 2022 Fire Code apply to 2012?", "gt": "No"},
    {"id": 555, "q": "Does 2020 Electrical Code apply to 2015?", "gt": "No"},
    {"id": 556, "q": "Are Mechanical and Plumbing in same doc?", "gt": "No"},
    {"id": 557, "q": "Are Fire and Electrical in same doc?", "gt": "No"},
    {"id": 558, "q": "Is 'Specification' part of document title?", "gt": "Yes"},
    {"id": 559, "q": "Is 'Code of Practice' part of document title?", "gt": "Yes"},
    {"id": 560, "q": "Is 'Government Buildings' part of title?", "gt": "Yes"},

    # --- Group 9: Final Random Logic Checks ---
    {"id": 561, "q": "Is 2.5x limit for start current?", "gt": "Yes"},
    {"id": 562, "q": "Is 2.5x limit for pump mass?", "gt": "Yes"},
    {"id": 563, "q": "Is 2.5x limit for tray spacing?", "gt": "No"},
    {"id": 564, "q": "Is 50mm for HV label?", "gt": "Yes"},
    {"id": 565, "q": "Is 50mm for acoustic lag?", "gt": "Yes"},
    {"id": 566, "q": "Is 50mm for screw protrusion?", "gt": "No"},
    {"id": 567, "q": "Is 20mm for conduit?", "gt": "Yes"},
    {"id": 568, "q": "Is 20mm for tray clearance?", "gt": "Yes"},
    {"id": 569, "q": "Is 20mm for copper flare limit?", "gt": "Yes"},
    {"id": 570, "q": "Is 20mm for cable standard split?", "gt": "Yes"},
    {"id": 571, "q": "Is 1.2m for tray spacing?", "gt": "Yes"},
    {"id": 572, "q": "Is 1.5m for mesh tray spacing?", "gt": "Yes"},
    {"id": 573, "q": "Is 1.2m for HV label spacing?", "gt": "No (3m)"},
    {"id": 574, "q": "Is 100mm for GS screwed limit?", "gt": "Yes"},
    {"id": 575, "q": "Is 100mm for inertia thickness?", "gt": "Yes"},
    {"id": 576, "q": "Is 150mm for inertia projection?", "gt": "Yes"},
    {"id": 577, "q": "Is 150mm for saddle spacing?", "gt": "Yes"},
    {"id": 578, "q": "Is 150mm for GS flanged limit?", "gt": "Yes"},
    {"id": 579, "q": "Is 50mm for GS fire screwed limit?", "gt": "Yes"},
    {"id": 580, "q": "Is 50Hz for auto-transformer?", "gt": "Yes"},
    {"id": 581, "q": "Is 60W for lighting?", "gt": "Yes"},
    {"id": 582, "q": "Is 130g/m2 for zinc?", "gt": "Yes"},
    {"id": 583, "q": "Is 2400mm for fan height?", "gt": "Yes"},
    {"id": 584, "q": "Is 2600mm for fan height?", "gt": "Yes"},
    {"id": 585, "q": "Is 10mm for loop clearance?", "gt": "Yes"},
    {"id": 586, "q": "Is 10mm for bolt clearance?", "gt": "Yes"},
    {"id": 587, "q": "Is 30-35 deg for swing?", "gt": "Yes"},
    {"id": 588, "q": "Is 99.9% for zinc purity?", "gt": "Yes"},
    {"id": 589, "q": "Is 24kg/m3 for density?", "gt": "Yes"},
    {"id": 590, "q": "Is 5kg/m2 for lead weight?", "gt": "Yes"},
    {"id": 591, "q": "Is 0.5x for deflection?", "gt": "Yes"},
    {"id": 592, "q": "Is 45kV for impulse?", "gt": "Yes"},
    {"id": 593, "q": "Is 3.3kV for system?", "gt": "Yes"},
    {"id": 594, "q": "Is 6.6kV for system?", "gt": "Yes"},
    {"id": 595, "q": "Is 11kV for system?", "gt": "Yes"},
    {"id": 596, "q": "Is 25mm for seal projection?", "gt": "Yes"},
    {"id": 597, "q": "Is 1 min for CCMS?", "gt": "Yes"},
    {"id": 598, "q": "Is 5% for CCMS step?", "gt": "Yes"},
    {"id": 599, "q": "Is 95% for CCMS max?", "gt": "Yes"},
    {"id": 600, "q": "Is 120 min for CCMS max?", "gt": "Yes"}
]
# ================= 3. Auto Judge =================
def judge_answer(prompt, answer, gt):
    """
    Automatic grading logic:
    1. Refusal detection
    2. Trap detection (prevent cross-domain hallucination)
    3. Numeric/keyword matching
    """
    if not answer or not isinstance(answer, str): return "Error"
    
    p = str(prompt).lower()
    ans = str(answer).lower()
    g = str(gt).lower()
    
    # 0. Refusal handling
    refusal_words = ["no information", "not found", "cannot answer", "unable to determine"]
    if any(w in ans for w in refusal_words):
        # If ground truth also indicates unknown, count as correct
        if "no" in g and len(g) < 15: return "Correct (Refusal)"
        if "check" in g or "context" in g: return "Correct (Refusal)"
        return "Refusal" # otherwise counted as refusal

    # 1. Trap detection (Cross-Domain Trap) - strength of Graph RAG
    # If asking about Fire but answering Plumbing materials (Cement/Bitumen), mark incorrect
    if "fire" in p and ("cement" in ans or "bitumen" in ans): return "Incorrect (Wrong Spec)"
    # If asking about Drainage but answering Plumbing parameter (2.5x), mark incorrect
    if "drainage" in p and "2.5" in ans: return "Incorrect (Hallucination)"

    # 2. Exact numeric matching
    # Extract numbers from GT (e.g. 50, 1.2, 2.5)
    gt_nums = re.findall(r'\d+(?:\.\d+)?', g)
    ans_nums = re.findall(r'\d+(?:\.\d+)?', ans)
    if gt_nums:
        # If all numbers in GT appear in answer
        if all(n in ans_nums for n in gt_nums):
            return "Correct"
        else:
            # If Yes/No question, mismatch counts as incorrect
            if "yes" in g or "no" in g: return "Incorrect (Value Mismatch)"

    # 3. Yes/No logic
    if g.startswith("yes"):
        return "Correct" if "yes" in ans and "no" not in ans[:10] else "Incorrect"
    if g.startswith("no"):
        return "Correct" if "no" in ans and "yes" not in ans[:10] else "Incorrect"

    # 4. Keyword fuzzy matching (fallback)
    # Remove stopwords
    stops = ["the", "is", "of", "to", "in", "required", "standard", "a", "an"]
    gt_tokens = set([w for w in re.split(r'\W+', g) if w not in stops and len(w)>2])
    if not gt_tokens: return "Unsure"
    
    match_count = sum(1 for t in gt_tokens if t in ans)
    if match_count / len(gt_tokens) >= 0.6: # 60% keyword match
        return "Correct"
    
    return "Incorrect"

# ================= 4. Service Connections =================
print("🔌 Connecting to services...")
try:
    client = OpenAI(api_key=API_KEY, base_url=BASE_URL)
    driver = GraphDatabase.driver(NEO4J_URI, auth=NEO4J_AUTH)
    # Test connection
    with driver.session() as session: session.run("RETURN 1")
    print("✅ Services connected successfully")
except Exception as e:
    print(f"❌ Connection failed: {e}")
    exit()

# Load local text data
try:
    with open(TEXT_DATA_FILE, 'r', encoding='utf-8') as f:
        KB_ENTRIES = json.load(f).get('entries', [])
except:
    print(f"⚠️ Warning: {TEXT_DATA_FILE} not found, Text RAG will not work.")
    KB_ENTRIES = []

# --- Retrieval functions ---
def retrieve_text(query):
    # Simple keyword-based text retrieval
    if not KB_ENTRIES: return ""
    hits = []
    q_words = str(query).lower().split()
    for entry in KB_ENTRIES:
        content = str(entry.get('text', ''))
        score = sum(1 for w in q_words if w in content.lower())
        if score > 0: hits.append((score, content))
    hits.sort(key=lambda x:x[0], reverse=True)
    return "\n---\n".join([h[1] for h in hits[:3]])

def retrieve_graph(query):
    # V8 Ultimate retrieval
    try:
        with driver.session() as session:
            kws = [w for w in str(query).lower().replace("?", "").split() if len(w)>2]
            if not kws: kws = ["general"]
            cypher = f"""
            WITH {json.dumps(kws)} as kws
            MATCH (n) WHERE ANY(k IN keys(n) WHERE ANY(w IN kws WHERE toLower(toString(n[k])) CONTAINS w))
            RETURN DISTINCT labels(n)[0] as type, properties(n) as props LIMIT 10
            """
            res = session.run(cypher)
            texts = []
            for r in res:
                props = r['props']
                # Simplified output
                p_str = ", ".join([f"{k}: {v}" for k,v in props.items() if k not in ['id', 'embedding']])
                texts.append(f"[{r['type']}] {p_str}")
            return "\n".join(texts)
    except Exception as e: return f"Graph Error: {e}"

def call_llm(sys_prompt, user_prompt):
    try:
        resp = client.chat.completions.create(
            model="deepseek-chat",
            messages=[{"role":"system","content":sys_prompt}, {"role":"user","content":user_prompt}],
            temperature=0.1
        )
        return resp.choices[0].message.content
    except Exception as e: return f"Error: {e}"

# ================= 5. Main Loop (with resume support) =================
def run_night_mode():
    # 1. Check for existing progress
    done_ids = set()
    existing_data = []
    
    if os.path.exists(OUTPUT_FILE):
        print(f"📂 Existing result file found: {OUTPUT_FILE}, loading progress...")
        try:
            df_exist = pd.read_excel(OUTPUT_FILE)
            if 'ID' in df_exist.columns:
                done_ids = set(df_exist['ID'].tolist())
                existing_data = df_exist.to_dict('records')
            print(f"✅ Completed {len(done_ids)} questions, will skip them.")
        except:
            print("⚠️ Failed to read, will overwrite old file.")

    print(f"🏃‍♂️ Starting benchmark! Remaining {len(QUESTIONS_DATA) - len(done_ids)} questions.")
    
    current_batch = []
    
    for item in tqdm(QUESTIONS_DATA):
        if item['id'] in done_ids:
            continue # skip completed
            
        q = item['q']
        gt = item['gt']
        
        row = {
            "ID": item['id'],
            "Question": q,
            "Ground Truth": gt,
            "S_LLM_Ans": "", "S_LLM_Res": "",
            "Text_RAG_Ans": "", "Text_RAG_Res": "",
            "Graph_RAG_Ans": "", "Graph_RAG_Res": ""
        }
        
        # --- Model 1: Standalone ---
        row["S_LLM_Ans"] = call_llm("You are an engineer.", q)
        row["S_LLM_Res"] = judge_answer(q, row["S_LLM_Ans"], gt)
        
        # --- Model 2: Text RAG ---
        txt_ctx = retrieve_text(q)
        if txt_ctx:
            row["Text_RAG_Ans"] = call_llm(f"Answer strictly based on:\n{txt_ctx}", q)
        else:
            row["Text_RAG_Ans"] = "No relevant text found."
        row["Text_RAG_Res"] = judge_answer(q, row["Text_RAG_Ans"], gt)
        
        # --- Model 3: Graph RAG ---
        graph_ctx = retrieve_graph(q)
        if graph_ctx:
            sys_prompt = "You are an assistant using a Knowledge Graph. Answer based strictly on the context. If context says 'Fire', do not use 'Plumbing' rules."
            row["Graph_RAG_Ans"] = call_llm(f"{sys_prompt}\nContext:\n{graph_ctx}", q)
        else:
            row["Graph_RAG_Ans"] = "No graph nodes found."
        row["Graph_RAG_Res"] = judge_answer(q, row["Graph_RAG_Ans"], gt)
        
        # Add to result set
        existing_data.append(row)
        current_batch.append(row)
        
        # === Key: save every 5 questions ===
        if len(current_batch) >= 5:
            save_excel(existing_data)
            current_batch = [] # clear cache

    # Final save after completion
    save_excel(existing_data)
    print("🎉 Overnight benchmarking task completed!")

def save_excel(data):
    """Save and format Excel"""
    df = pd.DataFrame(data)
    # Adjust column order
    cols = ["ID", "Question", "Ground Truth", 
            "S_LLM_Res", "S_LLM_Ans", 
            "Text_RAG_Res", "Text_RAG_Ans", 
            "Graph_RAG_Res", "Graph_RAG_Ans"]
    # Ensure columns exist
    final_cols = [c for c in cols if c in df.columns]
    df = df[final_cols]
    
    try:
        df.to_excel(OUTPUT_FILE, index=False)
        
        # Formatting
        wb = load_workbook(OUTPUT_FILE)
        ws = wb.active
        
        # Color definitions
        green = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # light green
        red = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")   # light red
        yellow = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # light yellow (refusal)
        
        # Apply coloring to result columns
        res_cols = [4, 6, 8] 
        
        for r in range(2, ws.max_row + 1):
            for c_idx in res_cols:
                cell = ws.cell(row=r, column=c_idx)
                val = str(cell.value)
                if "Correct" in val:
                    cell.fill = green
                elif "Incorrect" in val:
                    cell.fill = red
                elif "Refusal" in val:
                    cell.fill = yellow
                    
        wb.save(OUTPUT_FILE)
    except Exception as e:
        print(f"⚠️ Error while saving (does not affect execution): {e}")

if __name__ == "__main__":
    run_night_mode()
